from flask import Flask, render_template
from openpyxl import load_workbook
import cv2
import numpy as np
import xlsxwriter

cap = cv2.VideoCapture(0)
_, img = cap.read()
fh, fw, _ = img.shape
img = cv2.resize(img, (fw, fh))

file = open("loop1.txt", "r")
roiss11 = np.loadtxt("loop1.txt", dtype=int)
file.close()
mask1 = np.zeros((fh, fw), dtype=np.uint8)
cv2.drawContours(mask1, [roiss11], -1, (255, 255, 255), -1, cv2.LINE_AA)
mask1 = cv2.resize(mask1, (fw, fh), cv2.INTER_NEAREST)

file = open("loop2.txt", "r")
roiss22 = np.loadtxt("loop2.txt", dtype=int)
file.close()
mask2 = np.zeros((fh, fw), dtype=np.uint8)
cv2.drawContours(mask2, [roiss22], -1, (255, 255, 255), -1, cv2.LINE_AA)
mask2 = cv2.resize(mask2, (fw, fh), cv2.INTER_NEAREST)

file = open("loop3.txt", "r")
roiss33 = np.loadtxt("loop3.txt", dtype=int)
file.close()
mask3 = np.zeros((fh, fw), dtype=np.uint8)
cv2.drawContours(mask3, [roiss33], -1, (255, 255, 255), -1, cv2.LINE_AA)
mask3 = cv2.resize(mask3, (fw, fh), cv2.INTER_NEAREST)

file = open("loop4.txt", "r")
roiss44 = np.loadtxt("loop4.txt", dtype=int)
file.close()
mask4 = np.zeros((fh, fw), dtype=np.uint8)
cv2.drawContours(mask4, [roiss44], -1, (255, 255, 255), -1, cv2.LINE_AA)
mask4 = cv2.resize(mask4, (fw, fh), cv2.INTER_NEAREST)
#cv2.imshow('mask3', mask3)

file = open("loop5.txt", "r")
roiss55 = np.loadtxt("loop5.txt", dtype=int)
file.close()
mask5 = np.zeros((fh, fw), dtype=np.uint8)
cv2.drawContours(mask5, [roiss55], -1, (255, 255, 255), -1, cv2.LINE_AA)
mask5 = cv2.resize(mask5, (fw, fh), cv2.INTER_NEAREST)
#cv2.imshow('mask3', mask3)

file = open("loop6.txt", "r")
roiss66 = np.loadtxt("loop6.txt", dtype=int)
file.close()
mask6 = np.zeros((fh, fw), dtype=np.uint8)
cv2.drawContours(mask6, [roiss66], -1, (255, 255, 255), -1, cv2.LINE_AA)
mask6 = cv2.resize(mask6, (fw, fh), cv2.INTER_NEAREST)

prototxt = 'MobileNetSSD_deploy.prototxt.txt'
model = 'MobileNetSSD_deploy.caffemodel'

CLASSES = ["background", "aeroplane", "bicycle", "bird", "boat",
           "bottle", "bus", "car", "cat", "chair", "cow", "diningtable",
           "dog", "horse", "motorbike", "person", "pottedplant", "sheep",
           "sofa", "train", "tvmonitor"]
COLORS = np.random.uniform(0, 255, size=(len(CLASSES), 3))
net = cv2.dnn.readNetFromCaffe(prototxt, model)

app = Flask(__name__)

seconds1 = 0
minutes1 = 0
hours1 = 0
days1 = 0
rate = 100
seconds2 = 0
minutes2 = 0
hours2 = 0
days2 = 0
seconds3 = 0
minutes3 = 0
hours3 = 0
days3 = 0
seconds4 = 0
minutes4 = 0
hours4 = 0
days4 = 0
seconds5 = 0
minutes5 = 0
hours5 = 0
days5 = 0
seconds6 = 0
minutes6 = 0
hours6 = 0
days6 = 0


@app.route('/')
def index():
    book1 = load_workbook("park1.xlsx")
    data1 = book1.active
    data1["A1"] = 'FREE'
    data1["B1"] = 'FREE'
    data1["C1"] = 'FREE'
    data1["A2"] = 'FREE'
    data1["B2"] = 'FREE'
    data1["C2"] = 'FREE'
    book1.save(filename="park1.xlsx")

    book2 = load_workbook("park2.xlsx")
    data2 = book2.active
    data2["A1"] = '00:00:00'
    data2["B1"] = '00:00:00'
    data2["C1"] = '00:00:00'
    data2["A2"] = '00:00:00'
    data2["B2"] = '00:00:00'
    data2["C2"] = '00:00:00'
    book2.save(filename="park2.xlsx")

    book3 = load_workbook("park3.xlsx")
    data3 = book3.active
    data3["A1"] = 'Rs.0.0'
    data3["B1"] = 'Rs.0.0'
    data3["C1"] = 'Rs.0.0'
    data3["A2"] = 'Rs.0.0'
    data3["B2"] = 'Rs.0.0'
    data3["C2"] = 'Rs.0.0'
    book3.save(filename="park3.xlsx")

    global cap, mask1, mask2, mask3, seconds1, minutes1, hours1, days1, seconds2, minutes2, hours2, days2
    global seconds3, minutes3, hours3, days3, seconds4, minutes4, hours4, days4, seconds5, minutes5, hours5, days5
    global seconds6, minutes6, hours6, days6
    _, frame = cap.read()
    (fH, fW) = frame.shape[:2]
    frame = cv2.resize(frame, (fW, fH))
    res1 = cv2.bitwise_and(frame, frame, mask=mask1)
    res2 = cv2.bitwise_and(frame, frame, mask=mask2)
    res3 = cv2.bitwise_and(frame, frame, mask=mask3)
    res4 = cv2.bitwise_and(frame, frame, mask=mask4)
    res5 = cv2.bitwise_and(frame, frame, mask=mask5)
    res6 = cv2.bitwise_and(frame, frame, mask=mask6)

    blob1 = cv2.dnn.blobFromImage(cv2.resize(res1, (300, 300)), 0.007843, (300, 300), 127.5)
    blob2 = cv2.dnn.blobFromImage(cv2.resize(res2, (300, 300)), 0.007843, (300, 300), 127.5)
    blob3 = cv2.dnn.blobFromImage(cv2.resize(res3, (300, 300)), 0.007843, (300, 300), 127.5)
    blob4 = cv2.dnn.blobFromImage(cv2.resize(res4, (300, 300)), 0.007843, (300, 300), 127.5)
    blob5 = cv2.dnn.blobFromImage(cv2.resize(res5, (300, 300)), 0.007843, (300, 300), 127.5)
    blob6 = cv2.dnn.blobFromImage(cv2.resize(res6, (300, 300)), 0.007843, (300, 300), 127.5)

    net.setInput(blob1)
    rate = 100
    detections1 = net.forward()
    if detections1 is not None:
        for i in np.arange(0, detections1.shape[2]):
            confidence = detections1[0, 0, i, 2]
            if confidence < 0.2:
                continue
            idx1 = int(detections1[0, 0, i, 1])
            dims = np.array([fW, fH, fW, fH])
            box = detections1[0, 0, i, 3:7] * dims
            (startX1, startY1, endX1, endY1) = box.astype("int")
            label = "{}: {:.2f}%".format(CLASSES[idx1], confidence * 100)
            y = startY1 - 15 if startY1 - 15 > 15 else startY1 + 15
            if CLASSES[idx1] == "car" or CLASSES[idx1] == "bus" or CLASSES[idx1] == "motorbike" or CLASSES[
                idx1] == "bottle":
                data1["A1"] = 'OCCUPIED'
                book1.save(filename="park1.xlsx")

                workbook = load_workbook(filename="park2.xlsx")
                sheet = workbook.active
                seconds1 = seconds1 + 1
                if seconds1 % 61 == 0:
                    minutes1 = minutes1 + 1
                    seconds1 = 1
                    if minutes1 % 61 == 0:
                        hours1 = hours1 + 1
                        minutes1 = 1
                        if hours1 % 25 == 0:
                            days1 = days1 + 1
                            hours1 = 1
                sheet['A1'] = str(days1).rjust(2, '0') + ' days,' + str(hours1).rjust(2, '0') + ':' + str(
                    minutes1).rjust(2, '0') + ':' + str(seconds1).rjust(2, '0')
                workbook.save(filename="park2.xlsx")
                data2 = sheet

                fees = (rate * hours1) + rate
                workbook = load_workbook(filename="park3.xlsx")
                sheet = workbook.active
                sheet['A1'] = 'Rs.' + str(fees)
                workbook.save(filename="park3.xlsx")
                data3 = sheet

            else:
                print('no vehicle in A1...')
                seconds1 = 0
                minutes1 = 0
                hours1 = 0
                days1 = 0


    net.setInput(blob2)
    detections2 = net.forward()
    if detections2 is not None:
        for i in np.arange(0, detections2.shape[2]):
            confidence = detections2[0, 0, i, 2]
            if confidence < 0.2:
                continue
            idx2 = int(detections2[0, 0, i, 1])
            dims = np.array([fW, fH, fW, fH])
            box = detections2[0, 0, i, 3:7] * dims
            (startX2, startY2, endX2, endY2) = box.astype("int")
            label = "{}: {:.2f}%".format(CLASSES[idx2], confidence * 100)
            y = startY2 - 15 if startY2 - 15 > 15 else startY2 + 15
            if CLASSES[idx2] == "car" or CLASSES[idx2] == "bus" or CLASSES[idx2] == "motorbike":
                data1["B1"] = 'OCCUPIED'
                book1.save(filename="park1.xlsx")

                workbook = load_workbook(filename="park2.xlsx")
                sheet = workbook.active
                seconds2 = seconds2 + 1
                if seconds2 % 61 == 0:
                    minutes2 = minutes2 + 1
                    seconds2 = 1
                    if minutes2 % 61 == 0:
                        hours2 = hours2 + 1
                        minutes2 = 1
                        if hours2 % 25 == 0:
                            days2 = days2 + 1
                            hours2 = 1
                sheet['B1'] = str(days2).rjust(2, '0') + ' days,' + str(hours2).rjust(2, '0') + ':' + str(
                    minutes2).rjust(2, '0') + ':' + str(seconds2).rjust(2, '0')
                workbook.save(filename="park2.xlsx")
                data2 = sheet

                fees = (rate * hours2) + rate
                workbook = load_workbook(filename="park3.xlsx")
                sheet = workbook.active
                sheet['B1'] = 'Rs.' + str(fees)
                workbook.save(filename="park3.xlsx")
                data3 = sheet

            else:
                print('no vehicle in B1...')
                seconds2 = 0
                minutes2 = 0
                hours2 = 0
                days2 = 0


    net.setInput(blob3)
    detections3 = net.forward()
    if detections3 is not None:
        for i in np.arange(0, detections3.shape[2]):
            confidence = detections3[0, 0, i, 2]
            if confidence < 0.2:
                continue
            idx3 = int(detections3[0, 0, i, 1])
            dims = np.array([fW, fH, fW, fH])
            box = detections3[0, 0, i, 3:7] * dims
            (startX3, startY3, endX3, endY3) = box.astype("int")
            label = "{}: {:.2f}%".format(CLASSES[idx3], confidence * 100)
            y = startY3 - 15 if startY3 - 15 > 15 else startY3 + 15
            if CLASSES[idx3] == "car" or CLASSES[idx3] == "bus" or CLASSES[idx3] == "motorbike" or CLASSES[
                idx3] == "bottle":
                data1["C1"] = 'OCCUPIED'
                book1.save(filename="park1.xlsx")

                workbook = load_workbook(filename="park2.xlsx")
                sheet = workbook.active
                seconds3 = seconds3 + 1
                if seconds3 % 61 == 0:
                    minutes3 = minutes3 + 1
                    seconds3 = 1
                    if minutes3 % 61 == 0:
                        hours3 = hours3 + 1
                        minutes3 = 1
                        if hours3 % 25 == 0:
                            days3 = days3 + 1
                            hours3 = 1
                sheet['C1'] = str(days3).rjust(2, '0') + ' days,' + str(hours3).rjust(2, '0') + ':' + str(
                    minutes3).rjust(2, '0') + ':' + str(seconds3).rjust(2, '0')
                workbook.save(filename="park2.xlsx")
                data2 = sheet

                fees = (rate * hours3) + rate
                workbook = load_workbook(filename="park3.xlsx")
                sheet = workbook.active
                sheet['C1'] = 'Rs.' + str(fees)
                workbook.save(filename="park3.xlsx")
                data3 = sheet

            else:
                print('no vehicle in C1...')
                seconds3 = 0
                minutes3 = 0
                hours3 = 0
                days3 = 0


    net.setInput(blob4)
    detections4 = net.forward()
    if detections4 is not None:
        for i in np.arange(0, detections4.shape[2]):
            confidence = detections4[0, 0, i, 2]
            if confidence < 0.2:
                continue
            idx4 = int(detections4[0, 0, i, 1])
            dims = np.array([fW, fH, fW, fH])
            box = detections4[0, 0, i, 3:7] * dims
            (startX4, startY4, endX4, endY4) = box.astype("int")
            label = "{}: {:.2f}%".format(CLASSES[idx4], confidence * 100)
            y = startY4 - 15 if startY4 - 15 > 15 else startY4 + 15
            if CLASSES[idx4] == "car" or CLASSES[idx4] == "bus" or CLASSES[idx4] == "motorbike" or CLASSES[
                idx4] == "bottle":
                data1["A2"] = 'OCCUPIED'
                book1.save(filename="park1.xlsx")

                workbook = load_workbook(filename="park2.xlsx")
                sheet = workbook.active
                seconds4 = seconds4 + 1
                if seconds4 % 61 == 0:
                    minutes4 = minutes4 + 1
                    seconds4 = 1
                    if minutes4 % 61 == 0:
                        hours4 = hours4 + 1
                        minutes4 = 1
                        if hours4 % 25 == 0:
                            days4 = days4 + 1
                            hours4 = 1
                sheet['A2'] = str(days4).rjust(2, '0') + ' days,' + str(hours4).rjust(2, '0') + ':' + str(
                    minutes4).rjust(2, '0') + ':' + str(seconds4).rjust(2, '0')
                workbook.save(filename="park2.xlsx")
                data2 = sheet

                fees = (rate * hours4) + rate
                workbook = load_workbook(filename="park3.xlsx")
                sheet = workbook.active
                sheet['A2'] = 'Rs.' + str(fees)
                workbook.save(filename="park3.xlsx")
                data3 = sheet

            else:
                print('no vehicle in A2...')
                seconds4 = 0
                minutes4 = 0
                hours4 = 0
                days4 = 0


    net.setInput(blob5)
    detections5 = net.forward()
    if detections5 is not None:
        for i in np.arange(0, detections5.shape[2]):
            confidence = detections5[0, 0, i, 2]
            if confidence < 0.2:
                continue
            idx5 = int(detections5[0, 0, i, 1])
            dims = np.array([fW, fH, fW, fH])
            box = detections5[0, 0, i, 3:7] * dims
            (startX5, startY5, endX5, endY5) = box.astype("int")
            label = "{}: {:.2f}%".format(CLASSES[idx5], confidence * 100)
            y = startY5 - 15 if startY5 - 15 > 15 else startY5 + 15
            if CLASSES[idx5] == "car" or CLASSES[idx5] == "bus" or CLASSES[idx5] == "motorbike" or CLASSES[
                idx5] == "bottle":
                data1["B2"] = 'OCCUPIED'
                book1.save(filename="park1.xlsx")

                workbook = load_workbook(filename="park2.xlsx")
                sheet = workbook.active
                seconds5 = seconds5 + 1
                if seconds5 % 61 == 0:
                    minutes5 = minutes5 + 1
                    seconds5 = 1
                    if minutes5 % 61 == 0:
                        hours5 = hours5 + 1
                        minutes5 = 1
                        if hours5 % 25 == 0:
                            days5 = days5 + 1
                            hours5 = 1
                sheet['B2'] = str(days5).rjust(2, '0') + ' days,' + str(hours5).rjust(2, '0') + ':' + str(
                    minutes5).rjust(2, '0') + ':' + str(seconds5).rjust(2, '0')
                workbook.save(filename="park2.xlsx")
                data2 = sheet

                fees = (rate * hours5) + rate
                workbook = load_workbook(filename="park3.xlsx")
                sheet = workbook.active
                sheet['B2'] = 'Rs.' + str(fees)
                workbook.save(filename="park3.xlsx")
                data3 = sheet

            else:
                print('no vehicle in B2...')
                seconds5 = 0
                minutes5 = 0
                hours5 = 0
                days5 = 0


    net.setInput(blob6)
    detections6 = net.forward()
    if detections6 is not None:
        for i in np.arange(0, detections6.shape[2]):
            confidence = detections6[0, 0, i, 2]
            if confidence < 0.2:
                continue
            idx6 = int(detections6[0, 0, i, 1])
            dims = np.array([fW, fH, fW, fH])
            box = detections6[0, 0, i, 3:7] * dims
            (startX6, startY6, endX6, endY6) = box.astype("int")
            label = "{}: {:.2f}%".format(CLASSES[idx6], confidence * 100)
            y = startY6 - 15 if startY6 - 15 > 15 else startY6 + 15
            if CLASSES[idx6] == "car" or CLASSES[idx6] == "bus" or CLASSES[idx6] == "motorbike" or CLASSES[
                idx6] == "bottle":
                data1["C2"] = 'OCCUPIED'
                book1.save(filename="park1.xlsx")

                workbook = load_workbook(filename="park2.xlsx")
                sheet = workbook.active
                seconds6 = seconds6 + 1
                if seconds6 % 61 == 0:
                    minutes6 = minutes6 + 1
                    seconds6 = 1
                    if minutes6 % 61 == 0:
                        hours6 = hours6 + 1
                        minutes6 = 1
                        if hours6 % 25 == 0:
                            days6 = days6 + 1
                            hours6 = 1
                sheet['C2'] = str(days6).rjust(2, '0') + ' days,' + str(hours6).rjust(2, '0') + ':' + str(
                    minutes6).rjust(2, '0') + ':' + str(seconds6).rjust(2, '0')
                workbook.save(filename="park2.xlsx")
                data2 = sheet

                fees = (rate * hours6) + rate
                workbook = load_workbook(filename="park3.xlsx")
                sheet = workbook.active
                sheet['C2'] = 'Rs.' + str(fees)
                workbook.save(filename="park3.xlsx")
                data3 = sheet

            else:
                print('no vehicle in C2...')
                seconds6 = 0
                minutes6 = 0
                hours6 = 0
                days6 = 0

    return render_template('park_table.html', sheet1=data1, sheet2=data2, sheet3=data3)


if __name__ == "__main__":
    app.run(host="localhost", port=int("5000"))
